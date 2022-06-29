from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from easy_api.service.data_file import get_datas_from_xlsx, get_xlsx_from_datas


@pytest.mark.parametrize(["datas", "fields", "expect", "message"], (
        ([
             ("name", "age"),
             ("Duo", "18"),
             ("Bing", "19"),
         ], "", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "basic"),
        ([
             ("名称", "年龄"),
             ("Duo", "18"),
             ("Bing", "19"),
         ], "name:名称,age:年龄", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "fields"),
        ([
             ("名称 ", " 年龄"),
             ("Duo", "18"),
             ("Bing", "19"),
         ], "name:名称,age:年龄", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "fields with bad field"),
))
def test_get_data_from_xlsx(datas, fields, expect, message):
    mock_file = BytesIO()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    for row in datas:
        ws.append(row)
    wb.close()
    wb.save(mock_file)

    mock_file.seek(0)
    file_datas = mock_file.read()
    result = get_datas_from_xlsx(file_datas, fields)
    assert result == expect, message


@pytest.mark.parametrize(["expect", "fields", "datas", "message"], (
        ([
             ["name", "age"],
             ["Duo", "18"],
             ["Bing", "19"],
         ], "", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "basic"),
        ([
             ["名称", "年龄"],
             ["Duo", "18"],
             ["Bing", "19"],
         ], "name:名称,age:年龄", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "fields"),
))
def test_get_xlsx_from_datas(expect, fields, datas, message):
    mock_file = BytesIO()
    get_xlsx_from_datas(mock_file, datas, fields)
    mock_file.seek(0)

    values = []
    ws = load_workbook(mock_file, read_only=True).active
    for i, row in enumerate(ws.iter_rows()):
        values.append([x.value for x in row])

    assert values == expect, message


class MockWrite:
    _file = None
    _finished = False

    def __init__(self):
        self._file = BytesIO()

    def write(self, chunk):
        if self._finished:
            raise Exception("the write file is finished")
        self._file.write(chunk)
        return len(chunk)

    def flush(self):
        self._finished = True


@pytest.mark.parametrize(["expect", "fields", "datas", "message"], (
        ([
             ["name", "age"],
             ["Duo", "18"],
             ["Bing", "19"],
         ], "", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "basic"),
        ([
             ["名称", "年龄"],
             ["Duo", "18"],
             ["Bing", "19"],
         ], "name:名称,age:年龄", [{"name": "Duo", "age": "18"}, {"name": "Bing", "age": "19"}], "fields"),
))
def test_get_xlsx_from_datas_with_tornado_wrote_method(expect, fields, datas, message):
    mock_write_file = MockWrite()
    get_xlsx_from_datas(mock_write_file, datas, fields)
    mock_write_file._file.seek(0)

    mock_read_file = BytesIO()
    mock_read_file.write(mock_write_file._file.read())
    mock_read_file.seek(0)

    values = []
    ws = load_workbook(mock_read_file, read_only=True).active
    for i, row in enumerate(ws.iter_rows()):
        values.append([x.value for x in row])

    assert values == expect, message
