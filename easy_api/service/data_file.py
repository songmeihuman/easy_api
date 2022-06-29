import os.path
from io import BytesIO
from typing import Protocol, runtime_checkable

from openpyxl import load_workbook, Workbook


@runtime_checkable
class OnlyWrite(Protocol):
    def write(self, chunk) -> int:
        pass

    def flush(self):
        pass


def get_datas_from_xlsx(file_data: bytes, fields: str = "") -> list:
    input_file = BytesIO()
    input_file.write(file_data)
    ws = load_workbook(input_file, read_only=True).active

    field_dict = {}
    if fields:
        for item in fields.split(','):
            k, v = item.split(':')
            field_dict[v] = k

    values = []
    header = []
    for i, row in enumerate(ws.iter_rows()):
        if i == 0:
            header = [field_dict.get(x.value.strip(), x.value) for x in row]
        else:
            values.append(dict(zip(header, [x.value for x in row])))

    return values


def get_datas_from_file(file_name: str, file_data: bytes, fields: str = ""):
    file_ext = os.path.splitext(file_name)[-1]
    if file_ext == '.xlsx':
        return get_datas_from_xlsx(file_data, fields)


def get_xlsx_from_datas(fp: OnlyWrite, datas: list, fields: str = ""):
    if not fields:
        header_dict = {x: x for x in datas[0]}
    else:
        header_dict = {}
        for h in fields.split(","):
            hs = h.split(':')
            header_dict[hs[0]] = hs[-1]

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet()

    # write header
    worksheet.append(tuple(header_dict.values()))

    # write remain datas
    for i, item in enumerate(datas, 2):
        values = [
            '%s' % item.get(x, "") for x in header_dict
        ]
        worksheet.append(values)

    workbook.close()
    workbook.save(fp)
